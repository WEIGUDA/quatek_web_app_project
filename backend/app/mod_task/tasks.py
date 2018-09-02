import datetime
import logging
import os
import re
import socketserver
import sys
import threading
import time
import smtplib
import io
from email.message import EmailMessage
from logging import handlers

from celery import Celery
from pymongo import MongoClient, UpdateOne, bulk
from openpyxl import Workbook
from sqlalchemy import inspect, MetaData, create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.engine.url import URL

from app.mod_gate.schema import Log, Base

# load configs
from instance.config_default import *

try:
    from instance.config_dev import *
except:
    pass

try:
    from instance.config_pro import *
except:
    pass

# logging
logger = logging.getLogger(os.path.splitext(
    os.path.split((os.path.abspath(__file__)))[1])[0])
logger.setLevel(logging.INFO)
logFormatter = logging.Formatter('%(asctime)s - %(message)s')
fileHandler = handlers.RotatingFileHandler(
    '{}{}'.format(os.path.splitext(os.path.abspath(__file__))[0], '.log'),
    maxBytes=1024 * 1024 * 10,  # 10MB
    backupCount=10,
)
fileHandler.setFormatter(logFormatter)
consoleHandler = logging.StreamHandler(sys.stdout)
consoleHandler.setFormatter(logFormatter)
logger.addHandler(fileHandler)
# logger.addHandler(consoleHandler)


# celery
app = Celery('quatek-task', broker=REDIS_URL, result_backend=REDIS_URL)
app.conf.update({
    'CELERY_MONGODB_SCHEDULER_DB': MONGODB_DB,
    'CELERY_MONGODB_SCHEDULER_COLLECTION': "schedules",
    'CELERY_MONGODB_SCHEDULER_URL': f"mongodb://{MONGODB_HOST}:{MONGODB_PORT}"
})


class UploadAllCardsHandler(socketserver.BaseRequestHandler):
    ''' send: b'\rSET CARD;0;00CF1974;5454;Gorden;RD Team;1;3;0000;Remark\n'
        recv: b'\rCARD 0;00CF1974;5454;Gorden;RD Team;1;3;0000;Remark\n'
    '''

    def handle(self):
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        gates = db.gate
        cardtests = db.card_test
        users = db.user
        system_config = db.system_config

        logger.info('start an UploadAllCardsHandler for {}'.format(self.client_address))
        self.request.settimeout(5)
        mc_client = {}
        # get mc from database
        try:
            self.request.sendall(b'GET MCID\r\n')
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            mc_client_id = data.replace('\r', '').replace('\n', '').split(' ')[1]
            mc_client = gates.find({'mc_id': mc_client_id})[0]
            gates.update_one({'mc_id': 'mc_client.mc_id'},
                             {'$set': {'ip': self.client_address[0], 'port': self.client_address[1]}})

            if 'MCID' not in data:
                raise Exception('get mc error, mc: {} {}'.format(mc_client, self.client_address))
        except:
            logger.exception('error in UploadAllCardsHandler')

        # set datetime for mc
        try:
            dt = datetime.datetime.utcnow()
            self.request.sendall('SET DATE {}\r\n'.format(
                dt.strftime('%Y-%m-%d')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'DATE' not in data:
                raise Exception('set date error for  {}'.format(self.client_address))

            self.request.sendall('SET TIME {}\r\n'.format(dt.strftime('%H:%M:%S')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'TIME' not in data:
                raise Exception('set time error for {}'.format(self.client_address))

        except:
            logger.exception('error in UploadAllCardsHandler')

        # add cards to mc
        try:
            for card in cards.find():
                belong_to_mc = card['belong_to_mc']
                # add to all mc
                if belong_to_mc == 'all' or not belong_to_mc:
                    command = 'SET CARD {card_counter},{card_number},{job_number},{name},{department},{gender},{card_category},0,{note}\r\n'.format(
                        **card).encode(encoding='GB18030')
                    logger.info(command.decode(encoding='GB18030'))
                    self.request.sendall(command)
                    data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
                    if 'CARD' not in data:
                        raise Exception('upload card error, card: {}, from {}'.format(
                            card, self.client_address))

                # add to configed mc
                else:
                    # [{'mc1': '0'}, {'mc2': 1}]
                    belong_to_mc = [{item.split(':')[0]: item.split(':')[1]} for item in belong_to_mc.split('|')]

                    belong_to_mc_dict = {}  # {'mc1': '0', 'mc2': 1}

                    for item in belong_to_mc:
                        belong_to_mc_dict.update(item)

                    if mc_client['name'] in belong_to_mc_dict:
                        self.request.sendall(
                            'SET CARD {card_counter},{card_number},{job_number},{name},{department},{gender},{card_category},{0},{note}\r\n'.format(
                                belong_to_mc_dict[mc_client['name']], **card).encode(encoding='GB18030'))

                        data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
                        if 'CARD' not in data:
                            raise Exception('upload card error, card: {}, from {}'.format(
                                card, self.client_address))
        except:
            logger.exception('error in UploadAllCardsHandler')

        logger.info('stop the UploadAllCardsHandler for {}'.format(self.client_address))
        time.sleep(self.server.p_data['server_last_time'])


class UpdateACardHandler(socketserver.BaseRequestHandler):
    ''' send: b'\rSET CARD;0;00CF1974;5454;Gorden;RD Team;1;3;0000;Remark\n'
        recv: b'\rCARD 0;00CF1974;5454;Gorden;RD Team;1;3;0000;Remark\n'
        param: card
    '''

    def handle(self):
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        gates = db.gate
        cardtests = db.card_test
        users = db.user
        system_config = db.system_config
        self.request.settimeout(5)

        logger.info('start an UpdateACardHandler for {}'.format(self.client_address))
        mc_client = {}
        card = self.server.p_data['card']

        # get mc from database
        try:
            self.request.sendall(b'GET MCID\r\n')
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            mc_client_id = data.replace('\r', '').replace('\n', '').split(' ')[1]
            mc_client = gates.find({'mc_id': mc_client_id})[0]

            if 'MCID' not in data:
                raise Exception('get mc error, from {}'.format(self.client_address))
        except:
            logger.exception('error in UpdateACardHandler')

        # set datetime for mc
        try:
            dt = datetime.datetime.utcnow()
            self.request.sendall('SET DATE {}\r\n'.format(
                dt.strftime('%Y-%m-%d')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'DATE' not in data:
                raise Exception('set date error for  {}'.format(self.client_address))

            self.request.sendall('SET TIME {}\r\n'.format(dt.strftime('%H:%M:%S')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'TIME' not in data:
                raise Exception('set time error for {}'.format(self.client_address))

        except:
            logger.exception('error in UpdateACardHandler')

        try:
            belong_to_mc = card['belong_to_mc']

            # add to all mc
            if belong_to_mc == 'all' or not belong_to_mc:
                command = 'SET CARD {card_counter},{card_number},{job_number},{name},{department},{gender},{card_category},0,{note}\r\n'.format(
                    **card).encode(encoding='GB18030')
                logger.info(command.decode(encoding='GB18030'))
                self.request.sendall(command)

                data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
                if 'CARD' not in data:
                    raise Exception('upload the card to all mc error, card: {} from{}'.format(
                        card, self.client_address))

            # add to configed mc
            else:
                # [{'mc1': '0'}, {'mc2': 1}]
                belong_to_mc = [{item.split(':')[0]: item.split(':')[1]} for item in belong_to_mc.split('|')]

                belong_to_mc_dict = {}  # {'mc1': '0', 'mc2': 1}
                for item in belong_to_mc:
                    belong_to_mc_dict.update(item)

                if mc_client['name'] in belong_to_mc_dict:
                    self.request.sendall(
                        'SET CARD {card_counter},{card_number},{job_number},{name},{department},{gender},{card_category},{0},{note}\r\n'.format(
                            belong_to_mc_dict[mc_client['name']], **card).encode(encoding='GB18030'))
                    data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
                    if 'CARD' not in data:
                        raise Exception('upload the card to 1 mc error, card: {}, from {}'.format(
                            card, self.client_address))
                else:
                    self.request.sendall('CLR CARD {}\r\n'.format(card['card_counter']).encode(encoding='GB18030'))
                    data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
                    if 'CARD' not in data:
                        raise Exception('update with delete a card to mc error, card: {}, from {}'.format(
                            card, self.client_address))
        except:
            logger.exception('error in UpdateACardHandler')

        logger.info('stop the UpdateACardHandler for mc {} {}'.format(mc_client, self.client_address))
        time.sleep(self.server.p_data['server_last_time'])


class DeleteACardHandler(socketserver.BaseRequestHandler):
    ''' send: b'CLR CARD;0\r\n'
        recv: b'CARD 0;; ; ; ;0;0; ; \r\n'
    '''

    def handle(self):
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        gates = db.gate
        cardtests = db.card_test
        users = db.user
        system_config = db.system_config

        self.request.settimeout(5)
        logger.info('start an DeleteACardHandler for {}'.format(self.client_address))
        card = self.server.p_data['card']

        # set datetime for mc
        try:
            dt = datetime.datetime.utcnow()
            self.request.sendall('SET DATE {}\r\n'.format(
                dt.strftime('%Y-%m-%d')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'DATE' not in data:
                raise Exception('set date error for  {}'.format(self.client_address))

            self.request.sendall('SET TIME {}\r\n'.format(dt.strftime('%H:%M:%S')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'TIME' not in data:
                raise Exception('set time error for {}'.format(self.client_address))

        except:
            logger.exception('error in DeleteACardHandler')

        # delete card from mc
        try:
            self.request.sendall('CLR CARD {}\r\n'.format(card['card_counter']).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'CARD' not in data:
                raise Exception('delete a card from mc error, card: {}, from {}'.format(
                    card, self.client_address))
        except:
            logger.exception('error in DeleteACardHandler')

        logger.info('stop the DeleteACardHandler for {}'.format(self.client_address))
        time.sleep(self.server.p_data['server_last_time'])


class GetCardTestLogHandler(socketserver.BaseRequestHandler):
    ''' send: b'GET LOG\r\n'
        recv：b'LOG 4294967295;0;00CF1974;3;1;16;1900-01-01 00:00:00;1;0;FREE;FREE;FREE;NO\r\n'
    '''

    def handle(self):
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        gates = db.gate
        cardtests = db.card_test
        users = db.user
        system_config = db.system_config

        logger.info('start a new handler for {}'.format(self.client_address))
        self.request.settimeout(5)
        all_data = []
        mc_client = {}
        all_cardtests = []

        # get mc from database
        try:
            self.request.sendall(b'GET MCID\r\n')
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            mc_client_id = data.replace('\r', '').replace('\n', '').split(' ')[1]
            mc_client = gates.find({'mc_id': mc_client_id})[0]

            if 'MCID' not in data:
                raise Exception('get mc error, mc: {} {}'.format(mc_client, self.client_address))
        except:
            logger.exception('error in GetCardTestLogHandler')

        # set datetime for mc
        try:
            dt = datetime.datetime.utcnow()
            self.request.sendall('SET DATE {}\r\n'.format(
                dt.strftime('%Y-%m-%d')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'DATE' not in data:
                raise Exception('set date error for  {}'.format(self.client_address))

            self.request.sendall('SET TIME {}\r\n'.format(dt.strftime('%H:%M:%S')).encode(encoding='GB18030'))
            data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))
            if 'TIME' not in data:
                raise Exception('set time error for {}'.format(self.client_address))

        except:
            logger.exception('error in GetCardTestLogHandler')

        #  read all logs from mc

        while True:
            try:
                self.request.sendall(b'GET LOG\r\n')
                data = re.sub(r'CSN.*\r\n|\r|LOG ', '', self.request.recv(1024).decode(encoding='GB18030'))

                if '0,0,00000000,0,0,0,0,0,,,,,' in data:
                    logger.info('break, no logs in mc')
                    break

                if not data:
                    logger.info('break, no data from {} {}'.format(mc_client, self.client_address))
                    break

                all_data.append(data)
                self.request.sendall('CLR LOG {}\r\n'.format(data.split(
                    ',')[0].replace('LOG ', '')).encode(encoding='GB18030'))
                self.request.recv(1024)

            except TimeoutError:
                logger.exception('except break, timeout from {} {}'.format(mc_client, self.client_address))
                break
            except:
                logger.exception('except break, timeout from {} {}'.format(mc_client, self.client_address))
                break

        # process data and save to database
        if all_data:
            try:
                all_data = ''.join(all_data)
                all_data = re.sub(r'CSN.*\r\n|\r|LOG ', '', all_data).split('\n')

                if not all_data[-1]:
                    all_data.pop()

                for data in all_data:
                    data = data.split(',')
                    temp_dict = {
                        'log_id': data[0],
                        'card_counter': data[1],
                        'card_number': data[2],
                        'card_category': data[3],
                        'in_out_symbol': data[4],
                        'mc_id': data[5],
                        'test_result': data[6],
                        'is_tested': data[7],
                        'RSG': data[8],
                        'hand': data[9],
                        'left_foot': data[10],
                        'right_foot': data[11],
                        'after_erg': data[12],
                    }

                    all_cardtests.append(temp_dict)

                for cardtest in all_cardtests:
                    cardtest['test_datetime'] = datetime.datetime.fromtimestamp(
                        int(cardtest['log_id']), tz=datetime.timezone.utc)

                # client = MongoClient(MONGODB_HOST, MONGODB_PORT)
                # db = client[MONGODB_DB]
                # cardtests = db.cardtest
                cardtests.insert_many(all_cardtests)
            except:
                logger.exception('error from: {} {}'.format(mc_client, self.client_address))

        logger.info('stop the GetCardTestLogHandler for {} {}'.format(mc_client, self.client_address))
        time.sleep(self.server.p_data['server_last_time'])


class DeleteAllCardsFromMcHandler(socketserver.BaseRequestHandler):
    def handle(self):
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        gates = db.gate
        cardtests = db.card_test
        users = db.user
        system_config = db.system_config

        # delete all cards from mc
        command_list = []
        try:
            for i in range(6000):
                command_list.append('CLR CARD {}\r\n'.format(i))

                self.request.sendall(''.join(command_list).encode(encoding='GB18030'))

        except:
            logger.exception('delete all cards error from {}'.format(self.client_address))


class ThreadedTCPServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
    def __init__(self, server_address, RequestHandlerClass, bind_and_activate=True, p_data=None):
        super().__init__(server_address, RequestHandlerClass, bind_and_activate=True)
        self.p_data = p_data

    timeout = 5
    allow_reuse_address = True


@app.task()
def update_all_cards_to_mc_task(host=SOCKET_HOST, port=SOCKET_PORT, server_last_time=1):
    server = ThreadedTCPServer((host, port), UploadAllCardsHandler, p_data={'server_last_time': server_last_time})
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.daemon = True
    server_thread.start()
    logger.info("start an update_all_cards task")
    time.sleep(server_last_time)
    server.shutdown()
    server.server_close()
    logger.info("stop the update_all_cards task")


@app.task()
def update_a_card_to_all_mc_task(card_dict, server_last_time=1):
    server = ThreadedTCPServer((SOCKET_HOST, SOCKET_PORT), UpdateACardHandler,
                               p_data={'card': card_dict, 'server_last_time': server_last_time})
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.daemon = True
    server_thread.start()
    logger.info("start an update_a_card task")
    time.sleep(server_last_time)
    server.shutdown()
    server.server_close()
    logger.info("stop the update_a_card task")


@app.task()
def delete_a_card_from_mc_task(card_dict, server_last_time=1):
    server = ThreadedTCPServer((SOCKET_HOST, SOCKET_PORT), DeleteACardHandler,
                               p_data={'card': card_dict, 'server_last_time': server_last_time})
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.daemon = True
    server_thread.start()
    logger.info("start a delete_a_card task")
    time.sleep(server_last_time)
    server.shutdown()
    server.server_close()
    logger.info("stop the delete_a_card task")


@app.task()
def get_logs_from_mc_task(server_last_time=1):
    server = ThreadedTCPServer((SOCKET_HOST, SOCKET_PORT), GetCardTestLogHandler,
                               p_data={'server_last_time': server_last_time})
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.daemon = True
    server_thread.start()
    logger.info("start a get_logs_from_mc task")
    time.sleep(server_last_time)
    server.shutdown()
    server.server_close()
    logger.info("stop the get_logs_from_mc task")


@app.task()
def delete_all_cards_task(server_last_time=1):
    server = ThreadedTCPServer((SOCKET_HOST, SOCKET_PORT), DeleteAllCardsFromMcHandler,
                               p_data={'server_last_time': server_last_time})
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.daemon = True
    server_thread.start()
    logger.info("start a get_logs_from_mc task")
    time.sleep(server_last_time)
    server.shutdown()
    server.server_close()
    logger.info("stop the get_logs_from_mc task")


@app.task()
def send_email_of_logs():
    # pymongo
    client = MongoClient(MONGODB_HOST, MONGODB_PORT)
    db = client[MONGODB_DB]
    cards = db.card
    gates = db.gate
    cardtests = db.card_test
    users = db.user
    system_config = db.system_config

    logger.info('start send_email_of_logs task')
    config = system_config.find()[0]
    logger.info(f'config: {config}')

    # process data in mongodb
    all_cards = cards.find()
    local_tz = datetime.datetime.now(datetime.timezone.utc).astimezone().tzinfo

    work_hours_start = datetime.datetime.now().replace(
        hour=int(config['work_hours'].split('-')[0].split(':')[0]),
        minute=int(config['work_hours'].split('-')[0].split(':')[1]),
        second=0,
        microsecond=0,
        tzinfo=local_tz,
    )
    work_hours_end = datetime.datetime.now().replace(
        hour=int(config['work_hours'].split('-')[1].split(':')[0]),
        minute=int(config['work_hours'].split('-')[1].split(':')[1]),
        second=0,
        microsecond=0,
        tzinfo=local_tz,
    )

    # 该测试而未测试
    tested_card_number_list = cardtests.find(
        {'test_datetime': {'$gte': work_hours_start, '$lte': work_hours_end}, }).distinct('card_number')

    all_cards_should_test_but_not_tested = cards.find({
        'card_number': {'$nin': tested_card_number_list}
    })
    wb = Workbook()
    ws_should_test_but_not_tested = wb.active
    ws_should_test_but_not_tested.title = "该测试而未测试"

    ws_should_test_but_not_tested.cell(row=1, column=1).value = '卡号号码'
    ws_should_test_but_not_tested.cell(row=1, column=2).value = '卡片类别'
    ws_should_test_but_not_tested.cell(row=1, column=3).value = '姓名'
    ws_should_test_but_not_tested.cell(row=1, column=4).value = '工号'
    ws_should_test_but_not_tested.cell(row=1, column=5).value = '部门'
    ws_should_test_but_not_tested.cell(row=1, column=6).value = '性别'
    ws_should_test_but_not_tested.cell(row=1, column=7).value = '其他说明'
    ws_should_test_but_not_tested.cell(row=1, column=8).value = '权限'
    ws_should_test_but_not_tested.cell(row=1, column=9).value = '卡号编号'

    for i, card in enumerate(all_cards_should_test_but_not_tested):
        ws_should_test_but_not_tested.cell(row=i+2, column=1).value = card.get('card_number', '')
        card_category = card.get('card_category', '')
        if card_category == '0':
            ws_should_test_but_not_tested.cell(row=i+2, column=2).value = 'VIP'
        elif card_category == '1':
            ws_should_test_but_not_tested.cell(row=i+2, column=2).value = '只测手'
        elif card_category == '2':
            ws_should_test_but_not_tested.cell(row=i+2, column=2).value = '只测脚'
        elif card_category == '3':
            ws_should_test_but_not_tested.cell(row=i+2, column=2).value = '手脚都测'
        ws_should_test_but_not_tested.cell(row=i+2, column=3).value = card.get('name', '')
        ws_should_test_but_not_tested.cell(row=i+2, column=4).value = card.get('job_number', '')
        ws_should_test_but_not_tested.cell(row=i+2, column=5).value = card.get('department', '')
        gender = card.get('gender', '')
        if gender == '0':
            ws_should_test_but_not_tested.cell(row=i+2, column=6).value = '女'
        elif gender == '1':
            ws_should_test_but_not_tested.cell(row=i+2, column=6).value = '男'
        ws_should_test_but_not_tested.cell(row=i+2, column=7).value = card.get('note', '')
        ws_should_test_but_not_tested.cell(row=i+2, column=8).value = card.get('belong_to_mc', '')
        ws_should_test_but_not_tested.cell(row=i+2, column=9).value = card.get('card_counter', '')

    # 已测试而未通过
    all_logs_tested_but_not_passed = cardtests.find({
        'test_result': '0',
        'test_datetime': {'$gte': work_hours_start, '$lte': work_hours_end},
    })

    ws_tested_but_not_passed = wb.create_sheet("已测试而未通过")
    ws_tested_but_not_passed.cell(row=1, column=1).value = '记录流水号'
    ws_tested_but_not_passed.cell(row=1, column=2).value = '卡片编号'
    ws_tested_but_not_passed.cell(row=1, column=3).value = '卡片号码'
    ws_tested_but_not_passed.cell(row=1, column=4).value = '卡片类型'
    ws_tested_but_not_passed.cell(row=1, column=5).value = '进出标志'
    ws_tested_but_not_passed.cell(row=1, column=6).value = '闸机 mc id'
    ws_tested_but_not_passed.cell(row=1, column=7).value = '测试时间'
    ws_tested_but_not_passed.cell(row=1, column=8).value = '是否通过'
    ws_tested_but_not_passed.cell(row=1, column=9).value = '是否测试'
    ws_tested_but_not_passed.cell(row=1, column=10).value = '手腕检测值'
    ws_tested_but_not_passed.cell(row=1, column=11).value = '左脚检测值'
    ws_tested_but_not_passed.cell(row=1, column=12).value = '右脚检测值'
    ws_tested_but_not_passed.cell(row=1, column=13).value = 'ERG后的值'
    ws_tested_but_not_passed.cell(row=1, column=14).value = 'RSG值'

    for i, log in enumerate(all_logs_tested_but_not_passed):
        ws_tested_but_not_passed.cell(row=i+2, column=1).value = log.get('log_id', '')
        ws_tested_but_not_passed.cell(row=i+2, column=2).value = log.get('card_counter', '')
        ws_tested_but_not_passed.cell(row=i+2, column=3).value = log.get('card_number', '')
        card_category = log.get('card_category', '')
        if card_category == '0':
            ws_tested_but_not_passed.cell(row=i+2, column=4).value = 'VIP'
        elif card_category == '1':
            ws_tested_but_not_passed.cell(row=i+2, column=4).value = '只测手'
        elif card_category == '2':
            ws_tested_but_not_passed.cell(row=i+2, column=4).value = '只测脚'
        elif card_category == '3':
            ws_tested_but_not_passed.cell(row=i+2, column=4).value = '手脚都测'
        else:
            ws_tested_but_not_passed.cell(row=i+2, column=4).value = card_category
        ws_tested_but_not_passed.cell(row=i+2, column=5).value = log.get('in_out_symbol', '')
        ws_tested_but_not_passed.cell(row=i+2, column=6).value = log.get('mc_id', '')
        ws_tested_but_not_passed.cell(row=i+2, column=7).value = log.get('test_datetime', '').astimezone(local_tz)
        test_result = log.get('test_result', '')
        if test_result == '0':
            ws_tested_but_not_passed.cell(row=i+2, column=8).value = '不通过'
        elif test_result == '1':
            ws_tested_but_not_passed.cell(row=i+2, column=8).value = '通过'
        is_tested = log.get('is_tested', '')
        if is_tested == '0':
            ws_tested_but_not_passed.cell(row=i+2, column=9).value = '不测试'
        elif is_tested == '1':
            ws_tested_but_not_passed.cell(row=i+2, column=9).value = '测试'
        ws_tested_but_not_passed.cell(row=i+2, column=10).value = log.get('hand', '')
        ws_tested_but_not_passed.cell(row=i+2, column=11).value = log.get('left_foot', '')
        ws_tested_but_not_passed.cell(row=i+2, column=12).value = log.get('right_foot', '')
        ws_tested_but_not_passed.cell(row=i+2, column=13).value = log.get('after_erg', '')
        ws_tested_but_not_passed.cell(row=i+2, column=14).value = log.get('rsg', '')

    # sending email
    file_name = f'report-{datetime.datetime.now()}'
    msg = EmailMessage()
    msg['Subject'] = file_name
    msg['From'] = config['smtp_username']
    msg['To'] = ', '.join(config['emails'].split(','))
    msg.preamble = file_name
    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    msg.add_attachment(excel_data.read(), maintype='application', subtype='vnd.ms-excel',
                       filename=f'{file_name}.xlsx')

    if config['smtp_use_ssl']:
        with smtplib.SMTP_SSL(config['smtp_host'], config['smtp_port']) as s:
            if config['smtp_use_tls']:
                s.starttls()
            s.login(config['smtp_username'], config['smtp_password'])
            s.send_message(msg)
    else:
        with smtplib.SMTP(config['smtp_host'], config['smtp_port']) as s:
            if config['smtp_use_tls']:
                s.starttls()
            s.login(config['smtp_username'], config['smtp_password'])
            s.send_message(msg)

    logger.info('stop send_email_of_logs task')


@app.task()
def save_to_other_database():
    logger.info('start save_to_other_database task')

    try:
        # pymongo
        client = MongoClient(MONGODB_HOST, MONGODB_PORT)
        db = client[MONGODB_DB]
        cards = db.card
        cardtests = db.card_test
        system_config = db.system_config
        config = system_config.find()[0]

        os.environ["NLS_LANG"] = 'AMERICAN_AMERICA.UTF8'
        # sqlalchemy
        engine = create_engine(
            URL(
                drivername=config['db_type'],
                host=config['db_host'],
                port=config['db_port'],
                database=config['db_name'],
                username=config['db_username'],
                password=config['db_password']
            ),
        )

        inspector = inspect(engine)

        if 'logs' not in inspector.get_table_names():
            logger.info('create logs table')
            Base.metadata.create_all(bind=engine)

        Session = sessionmaker(bind=engine)
        session = Session()

        logs = []
        uncopied_cardtests = list(cardtests.find({'is_copied_to_other_database': False}))

        all_cards = list(cards.find())
        for cardtest in uncopied_cardtests:
            users_list = [user for user in all_cards if user['card_number'] == cardtest['card_number']]
            if users_list:
                user = users_list[0]
            else:
                user = {}
            log = Log(
                id=str(cardtest.get('_id', '')),
                log_id=cardtest.get('log_id', ''),
                card_counter=cardtest.get('card_counter', ''),
                card_number=cardtest.get('card_number', ''),
                card_category=cardtest.get('card_category', ''),
                in_out_symbol=cardtest.get('in_out_symbol', ''),
                mc_id=cardtest.get('mc_id', ''),
                test_datetime=cardtest.get('test_datetime', datetime.datetime.utcnow()),
                test_result=cardtest.get('test_result', ''),
                is_tested=cardtest.get('is_tested', ''),
                hand=cardtest.get('hand', ''),
                left_foot=cardtest.get('left_foot', ''),
                right_foot=cardtest.get('right_foot', ''),
                after_erg=cardtest.get('after_erg', ''),
                rsg=cardtest.get('rsg', ''),
                name=user.get('name', ''),
                job_number=user.get('job_number', ''),
                department=user.get('department', ''),
                gender=user.get('gender', ''),
                note=user.get('note', ''),
                belong_to_mc=user.get('belong_to_mc', ''),
            )

            if log.card_category == '0':
                log.card_category = 'vip'
            if log.card_category == '1':
                log.card_category = '只测手'
            if log.card_category == '2':
                log.card_category = '只测脚'
            if log.card_category == '3':
                log.card_category = '手脚都测'
            if log.gender == '0':
                log.gender = '女'
            if log.gender == '1':
                log.gender = '男'
            if log.test_result == '0':
                log.test_result = '不通过'
            if log.test_result == '1':
                log.test_result = '通过'
            if log.is_tested == '0':
                log.is_tested = '不测试'
            if log.is_tested == '1':
                log.is_tested = '测试'

            logs.append(log)

        logger.info(str(logs))
        session.add_all(logs)
        session.commit()

        requests = []
        for cardtest in uncopied_cardtests:
            requests.append(UpdateOne({'_id': cardtest['_id']}, {'$set': {'is_copied_to_other_database': True}}))
        try:
            cardtests.bulk_write(requests)
        except bulk.InvalidOperation:
            logger.info('pymongo bulk InvalidOperation: No operations to execute')

        session.close()
        engine.dispose()
    except Exception as e:
        logger.exception(e)

    logger.info('stop save_to_other_database task')
